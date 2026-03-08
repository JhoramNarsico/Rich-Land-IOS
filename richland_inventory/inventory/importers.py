# inventory/importers.py

import csv
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import StringIO
from openpyxl import load_workbook

from django.db import transaction
from django.utils import timezone

from .models import HydraulicSow, POSSale, CustomerPayment

def _read_file_to_dicts(file_obj):
    """
    Reads a CSV or XLSX file and returns a list of dictionaries.
    Normalizes headers to lowercase_with_underscores.
    """
    data = []
    file_name = file_obj.name.lower()
    
    try:
        if file_name.endswith('.csv'):
            # Use StringIO to handle in-memory file and utf-8-sig to handle BOM
            decoded_file = file_obj.read().decode('utf-8-sig').splitlines()
            reader = csv.DictReader(decoded_file)
            # Normalize headers
            reader.fieldnames = [name.strip().lower().replace(' ', '_') for name in reader.fieldnames]
            data = list(reader)
        elif file_name.endswith('.xlsx'):
            wb = load_workbook(file_obj, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows and len(rows) > 1:
                headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[0]]
                for row in rows[1:]:
                    # Create dict, converting None to empty string to mimic CSV behavior
                    row_dict = {headers[i]: (str(val) if val is not None else '') for i, val in enumerate(row) if i < len(headers)}
                    data.append(row_dict)
    except Exception as e:
        # Return an error that can be caught by the view
        raise ValueError(f"Could not read the file. It might be corrupted or in an unsupported format. Error: {e}")
        
    return data

def import_ledger_entries_from_file(file_obj, customer, user):
    """
    Processes an uploaded file (CSV or XLSX) to import ledger entries.
    Expected headers: date, reference, description, charge, payment
    Returns (success_count_charges, success_count_payments, errors_list)
    """
    try:
        data = _read_file_to_dicts(file_obj)
    except ValueError as e:
        return 0, 0, [str(e)]

    if not data:
        return 0, 0, ["File is empty or does not contain data rows."]

    count_charges = 0
    count_payments = 0
    errors = []

    try:
        with transaction.atomic():
            for i, row in enumerate(data, start=2): # Start at 2 for row number in file
                try:
                    # 1. Parse Date
                    date_val = row.get('date')
                    txn_date_obj = None
                    if isinstance(date_val, datetime):
                        txn_date_obj = date_val
                    elif date_val:
                        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y'):
                            try:
                                txn_date_obj = datetime.strptime(str(date_val).split(' ')[0], fmt)
                                break
                            except ValueError:
                                continue
                    
                    if not txn_date_obj:
                        errors.append(f"Row {i}: Invalid or missing date '{date_val}'. Use YYYY-MM-DD format.")
                        continue
                    
                    # Combine the date from the file with the current time of upload.
                    # This fixes the issue of all imported transactions being set to midnight.
                    # Use localtime to ensure the time of upload reflects the user's timezone, not UTC.
                    txn_date = datetime.combine(txn_date_obj.date(), timezone.localtime(timezone.now()).time())
                    txn_date = timezone.make_aware(txn_date) # Make it timezone-aware using settings.TIME_ZONE

                    ref = row.get('reference', '').strip()
                    desc = row.get('description', '').strip()

                    # 2. Parse Amounts
                    charge_val = Decimal(row.get('charge') or 0)
                    payment_val = Decimal(row.get('payment') or 0)

                    if payment_val < 0:
                        errors.append(f"Row {i}: Payment value cannot be negative.")
                        continue

                    if charge_val > 0:
                        errors.append(f"Row {i}: Creating new charges via import is disabled. Only payments are allowed.")
                        continue

                    # 3. Handle PAYMENT (Credit)
                    if payment_val > 0:
                        sale_to_pay = None
                        if ref:
                            # A reference was provided, so we MUST find a matching invoice.
                            sale_to_pay = POSSale.objects.filter(
                                customer=customer, 
                                receipt_id__iexact=ref
                            ).first()
                            
                            if not sale_to_pay:
                                # The reference was provided but it's invalid. This is an error.
                                errors.append(f"Row {i}: Reference '{ref}' does not match any existing invoice for this customer.")
                                continue # Skip to the next row

                        CustomerPayment.objects.create(
                            customer=customer, amount=payment_val, payment_date=txn_date,
                            reference_number=ref, notes=desc or f"Imported Payment from file.",
                            recorded_by=user, sale_paid=sale_to_pay
                        )
                        count_payments += 1

                except (InvalidOperation, ValueError) as e:
                    errors.append(f"Row {i}: Invalid number format for charge/payment. Details: {e}")
                
            if errors:
                raise ValueError("Errors found during processing. Rolling back changes.")

    except ValueError:
        return 0, 0, errors
    except Exception as e:
        return 0, 0, [f"An unexpected error occurred: {e}"]

    return count_charges, count_payments, []


def import_sow_from_file(file_obj, customer, user):
    """
    Processes an uploaded file (CSV or XLSX) to import SOW records.
    Returns (success_count, errors_list)
    """
    try:
        data = _read_file_to_dicts(file_obj)
    except ValueError as e:
        return 0, [str(e)]
        
    if not data:
        return 0, ["File is empty or does not contain data rows."]

    count = 0
    errors = []

    try:
        with transaction.atomic():
            for i, row in enumerate(data, start=2):
                try:
                    cost_val = row.get('cost', '0').strip()
                    cost = Decimal(cost_val) if cost_val else Decimal('0.00')

                    hose_type = row.get('hose_type', '').strip()
                    diameter = row.get('diameter', '').strip()
                    length = str(row.get('length', '')).strip()
                    fitting_a = row.get('fitting_a', '').strip()
                    fitting_b = row.get('fitting_b', '').strip()

                    missing = []
                    if not hose_type: missing.append("'Hose Type'")
                    if not diameter: missing.append("'Diameter'")
                    if not length: missing.append("'Length'")
                    if not fitting_a: missing.append("'Fitting A'")
                    if not fitting_b: missing.append("'Fitting B'")

                    if missing:
                        errors.append(f"Row {i}: Missing required field(s): {', '.join(missing)}.")
                        continue

                    HydraulicSow.objects.create(
                        customer=customer, created_by=user, hose_type=hose_type,
                        diameter=diameter, length=length or None,
                        pressure=row.get('pressure') or None, cost=cost if cost > 0 else None,
                        application=row.get('application', ''), fitting_a=fitting_a,
                        fitting_b=fitting_b, notes=row.get('notes', '')
                    )
                    count += 1
                except (InvalidOperation, ValueError) as e:
                    errors.append(f"Row {i}: Invalid number format for cost/length/pressure. Details: {e}")
            
            if errors:
                raise ValueError("Errors found during processing. Rolling back changes.")

    except ValueError:
        return 0, errors
    except Exception as e:
        return 0, [f"An unexpected error occurred: {e}"]

    return count, []